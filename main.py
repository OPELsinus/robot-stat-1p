import datetime
import os
import shutil
import time
from contextlib import suppress
from math import floor
from time import sleep
import pandas as pd

import psycopg2
from mouseinfo import screenshot
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from pywinauto import keyboard

from config import logger, tg_token, chat_id, db_host, robot_name, db_port, db_name, db_user, db_pass, ip_address, saving_path, download_path, ecp_paths, reports_saving_path
from tools.app import App
from tools.web import Web


def sql_create_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
        CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")} (
            started_time timestamp,
            ended_time timestamp,
            store_name text UNIQUE,
            executor_name text,
            status text,
            error_reason text,
            error_saved_path text,
            execution_time text,
            ecp_path text
            )
        '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def delete_by_id(id):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
                DELETE FROM ROBOT.{robot_name.replace("-", "_")} WHERE id = '{id}'
                '''
    c = conn.cursor()
    c.execute(table_create_query)
    conn.commit()
    c.close()
    conn.close()


def get_all_data():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            order by started_time asc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['started_time', 'ended_time', 'full_name', 'executor_name', 'status', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path']

    cur.close()
    conn.close()

    return df1


def get_data_by_name(store_name):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where store_name = '{store_name}'
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    # df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'status', 'error_reason', 'error_saved_path', 'execution_time']

    cur.close()
    conn.close()

    return len(df1)


def get_data_to_execute():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where (status != 'success' and status != 'processing')
            and (executor_name is NULL or executor_name = '{ip_address}')
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())

    with suppress(Exception):
        df1.columns = ['started_time', 'ended_time', 'full_name', 'executor_name', 'status', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path']

    cur.close()
    conn.close()

    return df1


def insert_data_in_db(started_time, store_name, executor_name, status_, error_reason, error_saved_path, execution_time, ecp_path_):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)

    print('Started inserting')
    # query_delete_id = f"""
    #         delete from ROBOT.{robot_name.replace("-", "_")}_2 where store_id = '{store_id}'
    #     """
    query_delete = f"""
        delete from ROBOT.{robot_name.replace("-", "_")} where store_name = '{store_name}'
    """
    query = f"""
        INSERT INTO ROBOT.{robot_name.replace("-", "_")} (started_time, ended_time, store_name, executor_name, status, error_reason, error_saved_path, execution_time, ecp_path)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    # ended_time = '' if status_ != 'success' else datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    ended_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    values = (
        started_time,
        ended_time,
        store_name,
        executor_name,
        status_,
        error_reason,
        error_saved_path,
        str(execution_time),
        ecp_path_
    )

    print(values)

    cursor = conn.cursor()

    cursor.execute(query_delete)
    # conn.autocommit = True
    try:
        cursor.execute(query_delete)
        # cursor.execute(query_delete_id)
    except Exception as e:
        print('GOVNO', e)
        pass
    try:
        cursor.execute(query, values)
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")

    conn.commit()

    cursor.close()
    conn.close()


def get_all_branches_with_codes():

    conn = psycopg2.connect(dbname='adb', host='172.16.10.22', port='5432',
                            user='rpa_robot', password='Qaz123123+')

    cur = conn.cursor(name='1583_first_part')

    query = f"""
        select db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name 
        from dwh_data.dim_branches db
        left join dwh_data.dim_store ds on db.id_sale_object = ds.sale_source_obj_id
        where ds.store_name like '%Торговый%' and current_date between ds.datestart and ds.dateend
        group by db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name
        order by ds.source_store_id
    """

    cur.execute(query)

    print('Executed')

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['branch_id', 'store_id', 'store_name', 'store_normal_name']

    cur.close()
    conn.close()

    return df1


def sign_ecp(ecp):
    logger.info('Started ECP')
    logger.info(f'KEY: {ecp}')
    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        if app.wait_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, timeout=30):
            app.find_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}).type_keys('Aa123456')

            sleep(2)
            keyboard.send_keys('{ENTER}')
            sleep(3)
            keyboard.send_keys('{ENTER}')
            app = None
            logger.info('Finished ECP')

        else:
            logger.info('Quit mazafaka1')
            app = None
            return 'broke'
    else:
        logger.info('Quit mazafaka')
        app = None
        return 'broke'


def save_screenshot(store):
    scr = screenshot()
    save_path = os.path.join(saving_path, 'Ошибки 1П')
    scr_path = str(os.path.join(os.path.join(saving_path, 'Ошибки 1П'), str(store + '.png')))
    scr.save(scr_path)

    return scr_path


def wait_loading(web, xpath):
    print('Started loading')
    ind = 0
    element = ''
    while True:
        try:
            print(web.get_element_display('//*[@id="loadmask-1315"]'))
            if web.get_element_display('//*[@id="loadmask-1315"]') == '':
                element = ''
            if (element == '' and web.get_element_display('//*[@id="loadmask-1315"]') == 'none') or (ind >= 500):
                print('Loaded')
                sleep(0.5)
                break
        except:
            print('No loader')
            break
        ind += 1
        sleep(0.05)


def send_file_to_tg(tg_token, chat_id, param, param1):
    pass


def create_and_send_final_report():
    df = get_all_data()

    df.columns = ['Время начала', 'Время окончания', 'Название филиала', 'Статус', 'Причина ошибки', 'Пусть сохранения скриншота', 'Время исполнения (сек)', 'Факт1', 'Факт2', 'Факт3', 'Сайт1', 'Сайт2', 'Сайт3']

    df['Время исполнения (сек)'] = df['Время исполнения (сек)'].astype(float)
    df['Время исполнения (сек)'] = df['Время исполнения (сек)'].round()

    df.to_excel('result.xlsx', index=False)

    workbook = load_workbook('result.xlsx')
    sheet = workbook.active

    red_fill = PatternFill(start_color="FFA864", end_color="FFA864", fill_type="solid")
    green_fill = PatternFill(start_color="A6FF64", end_color="A6FF64", fill_type="solid")

    for cell in sheet['D']:
        if cell.value == 'failed':
            cell.fill = red_fill
        if cell.value == 'success':
            cell.fill = green_fill

    for col in 'ABCDGH':

        max_length = max(len(str(cell.value)) for cell in sheet[col])

        if col == 'A' or col == 'B':
            max_length -= 3
        if col == 'D':
            max_length += 5
        if col == 'A':
            max_length -= 3

        sheet.column_dimensions[col].width = max_length

    for col in 'ABCDGEFGH':
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='center')

    workbook.save('result.xlsx')

    send_file_to_tg(tg_token, chat_id, 'Отправляем отчёт по заполнению', 'result.xlsx')


def wait_image_loaded(store_):
    found = False
    while True:
        for file in os.listdir(download_path):
            if '.jpg' in file and 'crdownload' not in file:
                shutil.move(os.path.join(download_path, file), os.path.join(reports_saving_path, store_.replace('_stat.xlsx', '') + '.jpg'))
                print(file)
                found = True
                break
        if found:
            break


def save_and_send(web, save):
    print('Saving and Sending')
    if save:
        web.execute_script_click_xpath("//span[text() = 'Сохранить']")
        sleep(1)
        print('Clicked Save')
        if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=5):
            web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")
    print('Clicking Send')
    errors_count = web.find_elements('//*[@id="statflc"]/ul/li/a')
    if len(errors_count) <= 1:
        print('ALL GOOD')
        web.execute_script_click_xpath("//span[text() = 'Отправить']")
        print('Clicked Send')
        web.wait_element("//input[@value = 'Персональный компьютер']", timeout=30)
        web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
    else:
        print('GOVNO OSHIBKA VYLEZLA')


def wait_loading_1p(web, store):

    for i in range(5):

        if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

        if web.wait_element("(//tr[@role='row'])[1]", timeout=5):

            if web.wait_element("//div[contains(text(), '1-П (кварт')]", timeout=3):
                web.find_element("//div[contains(text(), '1-П (кварт')]").click()

                return True

            # else:
            #     saved_path = save_screenshot(store.replace('_stat.xlsx', ''))
            #     web.close()
            #     web.quit()
            #
            #     print('Return those shit')
            #     return False

        else:
            web.driver.refresh()

    # saved_path = save_screenshot(store)
    # web.close()
    # web.quit()

    print('Return those shit')
    return False


def proverka_ecp(web):

    if web.wait_element('//*[@id="AgreeId_header_hd-textEl"]', timeout=.5):
        web.execute_script_click_xpath("//span[text() = 'Согласен']")


def start_single_branch(filepath, store, values_first_part, values_second_part):

    def pass_later():
        if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            print('PASSING LATER')
            web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

    print('Started web')

    ecp_auth = ''
    ecp_sign = ''
    for file in os.listdir(filepath):

        if 'AUTH' in file:
            ecp_auth = os.path.join(filepath, file)
        if 'GOST' in file:
            ecp_sign = os.path.join(filepath, file)

    print(ecp_auth, '|', ecp_sign)
    web = Web()
    web.run()
    web.get('https://cabinet.stat.gov.kz/')
    print('1.0')
    logger.info('Check-1')
    print('1.1')

    logger.info('refreshed')

    proverka_ecp(web=web)

    web.wait_element('//*[@id="idLogin"]')
    web.find_element('//*[@id="idLogin"]').click()
    print('1.2')
    proverka_ecp(web=web)

    # * --- deprecated (maybe useful in future)
    # web.wait_element('//*[@id="button-1077-btnEl"]')
    # web.find_element('//*[@id="button-1077-btnEl"]').click()
    # * ---
    # proverka_ecp(web=web)
    print()
    # web.wait_element('//*[@id="lawAlertCheck"]')
    # web.find_element('//*[@id="lawAlertCheck"]').click()
    web.execute_script_click_xpath("//input[@id='lawAlertCheck']")
    print('1.3')
    time.sleep(0.5)
    web.find_element('//*[@id="loginButton"]').click()

    logger.info('Check-2')

    time.sleep(1)

    # send_message_to_tg(tg_token, chat_id, f"Started ECP, {datetime.datetime.now()}")
    sign_ecp(ecp_auth)
    # send_message_to_tg(tg_token, chat_id, f"Finished ECP, {datetime.datetime.now()}")
    print('1.4')
    logged_in = web.wait_element("//a[text()='Выйти']", timeout=10)

    print('1.44')
    store = branch.split('\\')[-1]
    print('1.444')
    # sleep(1000)
    if logged_in:
        print('1.45')
        if web.find_element("//a[text()='Выйти']", timeout=30):
            print('1.5')

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                try:
                    print('1.6')
                    web.find_element("//span[contains(text(), 'Пройти позже')]", timeout=5).click()
                except:
                    save_screenshot(store)
                    # print('HUETA')
                    # sleep(200)
            logger.info('Check0')
            if web.wait_element('//*[@id="dontAgreeId-inputEl"]', timeout=5):
                web.find_element('//*[@id="dontAgreeId-inputEl"]').click()
                sleep(0.3)
                web.find_element('//*[@id="saveId-btnIconEl"]').click()
                sleep(1)

                # * --- Deprecated (maybe useful)
                # web.find_element('//*[@id="ext-gen1893"]').click()
                # web.find_element('//*[@id="boundlist-1327-listEl"]/ul/li').click()
                # * ---

                web.wait_element('//*[@id="keyCombo-inputEl"]')

                web.execute_script_click_xpath("//*[@id='keyCombo-inputEl']/../following-sibling::td//div")

                web.find_element("//li[contains(text(), 'Персональный компьютер')]").click()
                sleep(1.5)

                web.execute_script_click_xpath("//span[contains(text(), 'Продолжить')]")

                print('Done lol')
                sign_ecp(ecp_sign)
                print('Finished done lol')
                try:
                    if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                        web.find_element("//span[contains(text(), 'Пройти позже')]").click()

                except:
                    pass

            # web.wait_element('//*[@id="radio-1131-boxLabelEl"]')

            pass_later()
            print('OTCHETY')
            web.wait_element("//span[contains(text(), 'Мои отчёты')]")
            web.execute_script_click_xpath("//span[contains(text(), 'Мои отчёты')]")

            # ? Check if 1П exists

            pass_later()

            # * ------- Uncomment -------
            found_ = wait_loading_1p(web, store)

            # for _ in range(1):
            #
            #     is_loaded = True if len(web.find_elements("//div[contains(@class, 'x-grid-row-expander')]", timeout=15)) >= 1 else False
            #
            #     if is_loaded:
            #         if web.wait_element("//div[contains(text(), '1-П (кварт')]", timeout=3):
            #             web.find_element("//div[contains(text(), '1-П')]").click()
            #             found_ = True
            #
            #             web.find_element('//*[@id="createReportId-btnIconEl"]').click()
            #             break
            #
            #         else:
            #             saved_path = save_screenshot(store)
            #             web.close()
            #             web.quit()
            #
            #             print('Return those shit')
            #             return ['failed', saved_path, 'Нет 1-П']
            #
            #     else:
            #         web.refresh()

            if found_:
                web.find_element('//*[@id="createReportId-btnIconEl"]').click()

            if not found_:
                print('Calendar')
                web.find_element('//span[contains(text(), "Календарь")]').click()
                web.wait_element('//div[text() = "1-П"]')

                print('Waited')
                print(web.find_element('//div[text() = "1-П"]/../following-sibling::td[1]/div').get_attr('text'))
                if web.find_element('//div[text() = "1-П"]') and web.find_element('//div[text() = "1-П"]/../following-sibling::td[1]/div').get_attr('text') == 'квартал':
                    print('Here')
                    # web.execute_script_click_xpath('//div[text() = "1-П"]/../following-sibling::td//button/p')
                    web.find_element('//div[text() = "1-П"]/../following-sibling::td//button').click()

                # saved_path = save_screenshot(store)
                # web.close()
                # web.quit()
                #
                # print('Return those shit')
                # return ['failed', saved_path, 'Нет 1-П']

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

            sleep(1)

            # ? Switch to the second window
            web.driver.switch_to.window(web.driver.window_handles[-1])

            web.find_element('/html/body/div[1]').click()
            web.wait_element('//*[@id="td_select_period_level_1"]/span')
            web.execute_script_click_js("#btn-opendata")
            sleep(0.3)

            if not found_:
                if web.wait_element('//span[text() = "Подтвердите открытие формы."]', timeout=10):
                    web.execute_script_click_xpath("//span[text() = 'Подтвердите открытие формы.']/../..//span[text() = 'Открыть']")
                    sleep(4)
                    # web.execute_script_click_js("#btn-opendata")

            if web.get_element_display('/html/body/div[7]') == 'block':
                web.find_element('/html/body/div[7]/div[11]/div/button[2]').click()

                saved_path = save_screenshot(store)
                web.close()
                web.quit()

                print('Return that shit')
                return ['failed', saved_path, 'Выскочила ошиПочка']

            # logger.info('Check3')
            # sleep(1)
            web.wait_element('//*[@id="sel_statcode_accord"]/div/p/b[1]', timeout=100)
            web.execute_script_click_js("body > div:nth-child(16) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1) > span")
            # web.execute_script_click_xpath("//span[text() = 'Выбрать']")

            web.wait_element('//*[@id="sel_rep_accord"]/h3[1]/a')

            sites = []

            # ? Open new report to fill it
            # web.wait_element('//span[text() = "Выберите отчет"]')

            print('Clicking1')
            web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")
            # web.execute_script_click_xpath('/html/body/div[17]/div[11]/div/button[1]/span')

            # ? First page

            web.wait_element("//a[contains(text(), 'Страница 1')]", timeout=10)
            web.find_element("//a[contains(text(), 'Страница 1')]").click()
            print()
            id_ = 3
            for ind, key in enumerate(first.keys()):

                if key == 'Всего':
                    continue
                if first.get(key) > 0:
                    # print(key, first.get(key))
                    # print(f'//*[@id="{id_}"]/td[3]', f'//*[@id="{id_}_col_1"]')
                    web.find_element(f'//*[@id="3"]/td[2]').click()

                    web.find_element(f'//*[@id="{id_}_col_1"]').type_keys(str(key))

                    sleep(1.5)

                    keyboard.send_keys('{ENTER}')

                    # web.find_element(f'//*[@id="{ind + 1}"]/td[3]').click()
                    web.find_element(f'//*[@id="{id_}_col_2"]').type_keys(str(first.get(key)))

                    id_ += 1

            keyboard.send_keys('{TAB}')
            # sleep(100)
            # ? Second page
            web.wait_element("//a[contains(text(), 'Страница 2')]", timeout=10)
            web.find_element("//a[contains(text(), 'Страница 2')]").click()

            web.find_element('//*[@id="rtime"]').select('2')
            sleep(1)
            print('-----')

            id_ = 3
            for i in range(len(second)):

                cur_key = list(second.keys())[i]

                if cur_key == 'Всего':
                    continue

                web.find_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][2]").click()
                web.wait_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][2]//input")
                web.find_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][2]//input").type_keys(cur_key, delay=1)
                sleep(1)
                keyboard.send_keys('{ENTER}')
                print(cur_key)

                for ind, val in enumerate(second.get(cur_key)):

                    if val == 0 and ind >= 2:
                        continue
                    else:
                        web.find_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][{ind + 4}]").click(double=True)
                        print(second.get(cur_key)[ind])
                        # print(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][{ind + 4}]//input")
                        web.wait_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][{ind + 4}]//input")
                        web.find_element(f"//table[@id='tb_p1_e0']//tr[{id_}]/td[@role='gridcell'][{ind + 4}]//input").type_keys(str(second.get(cur_key)[ind]), delay=1)

                id_ += 1

            keyboard.send_keys('{TAB}')
            # ? Last page
            web.find_element("//a[contains(text(), 'Данные исполнителя')]").click()
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_2_0']", value='Естаева Акбота Канатовна')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_2_1']", value='7273391350')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_2_2']", value='7073882688')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_2_3']", value='Yestayeva@magnum.kz')

            save_and_send(web, save=True)
            # sleep(3000)
            sign_ecp(ecp_sign)
            # sleep(1000)

            wait_image_loaded(store)

            web.close()
            web.quit()

            print('Successed')
            return ['success', '', '']

            # return ['success', '', sites]

    else:

        saved_path = save_screenshot(store)

        web.close()
        web.quit()

        print('Srok istek')
        return ['failed', saved_path, 'Срок ЭЦП истёк']


def get_first_page(filepath):

    book = load_workbook(filepath, data_only=True)

    sheet = book['Стр 2-3']

    data_from_first_page = dict()

    data_from_first_page.update({'Всего': round(sheet['C12'].value)})
    data_from_first_page.update({'10110': round(sheet['C13'].value) + round(sheet['C14'].value)})
    data_from_first_page.update({'10120': round(sheet['C15'].value) + round(sheet['C16'].value)})
    for i in range(17, 24):
        try:
            if str(sheet[f'B{i}'].value)[:4] == '1089':
                data_from_first_page.update({str(sheet[f'B{i}'].value)[:4] + '1': round(sheet[f'C{i}'].value)})
            elif round(sheet[f'C{i}'].value) > 0:
                data_from_first_page.update({str(sheet[f'B{i}'].value)[:4] + '0': round(sheet[f'C{i}'].value)})
            # print(int(sheet[f'C{i}'].value))
        except:
            if str(sheet[f'B{i}'].value)[:4] == '1089':
                data_from_first_page.update({str(sheet[f'B{i}'].value)[:4] + '1': 0})
            else:
                data_from_first_page.update({str(sheet[f'B{i}'].value)[:4] + '0': 0})

    # for key, i in data_from_first_page.items():
    #     print(key, i)

    return data_from_first_page


def get_second_page(filepath):
    book = load_workbook(filepath, data_only=True)

    sheet = book['Стр 4-5']

    data_from_second_page = dict()

    data_from_second_page.update({'Всего': round(sheet['E49'].value)})
    print('KEK:', sheet['D46'].value)
    for i in range(6, 49):
        row = []
        if sheet[f'A{i}'].value is not None:
            for vals in 'DEFGHIJK':
                if sheet[f'{vals}{i}'].value is not None:
                    row.append(round(sheet[f'{vals}{i}'].value))
                else:
                    row.append(0)

            if sum(row) > 0:
                data_from_second_page.update({sheet[f'B{i}'].value: row})

    # for key, i in data_from_second_page.items():
    #     print(key, i)
    #
    # print(len(data_from_second_page))

    return data_from_second_page


def get_calculated_dicts(first_, second_):

    dict1 = first_.copy()
    dict2 = second_.copy()

    dick = dict({'Всего': dict2['Всего']})

    sum1, sum2 = 0, 0
    for key, val in first.items():
        if key != 'Всего':
            sum1 += val
            # print(str(key)[:-1], ',', val)
    # print('-------------')
    for key, val in second.items():
        if key != 'Всего':
            sum2 += val[1]
            # print(str(key)[:4], ',', val[1])
    print('SUMS:', sum1, sum2)
    # if sum1 != sum2: # dict1['Всего'] != dict2['Всего']

    for key, val in dict1.items():
        s = 0
        for key1, val1 in dict2.items():

            if key1 != 'Всего':
                if str(key)[:-1] == str(key1)[:4]:
                    s += dict2.get(key1)[1]

        dict1.update({key: s})
        # print(key, s)
    # for key, val in dict2.items():
    #     if key != 'Всего':
    #         # print(key, val)
    #         # print(str(key)[:4], sum(val))
    #         if dick.get(str(key)[:4] + '0') is None:
    #             if str(key)[:4] == '1089':
    #                 dick.update({str(key)[:4] + '1': val[1]})
    #             else:
    #                 dick.update({str(key)[:4] + '0': val[1]})
    #         else:
    #             if str(key)[:4] == '1089':
    #                 dick.update({str(key)[:4] + '1': val[1] + dick.get(str(key)[:4] + '1')})
    #             else:
    #                 dick.update({str(key)[:4] + '0': val[1] + dick.get(str(key)[:4] + '0')})
    # else:
    #     return dict1, dict2
    dict1.update({'Всего': dict2.get('Всего')})
    print('=======')
    print(dict1)
    print(dict2)
    print(dick)
    # print(sum(dick.values()) - dick['Всего'])
    # dick.pop('10610')
    # s, s1 = 0, 0
    # for key in dict1.keys():
    #
    #     if key == '10610' or key == 'Всего':
    #         continue
    #
    #     print(key, '|', dict1.get(key), dick.get(key), '|')
    #     dict1.update({key: dict1.get(key) - (dict1.get(key) - dick.get(key))})
    #     # print(key, '|', dict1.get(key), dick.get(key), '|', dict1.get(key) - dick.get(key))
    #     s += dict1.get(key)
    #     s1 += dick.get(key)
    #     # print('-----------------------------------------')
    #
    # # print('==========================')
    # dick['Всего'] = sum(dick.values()) - dick['Всего']
    # dict1['Всего'] = sum(dict1.values()) - dict1['Всего']
    # # print(s, s1)
    # # print(sum(dick.values()) - dick['Всего'], dick['Всего'])
    # # print(sum(dict1.values()) - dict1['Всего'], dict1['Всего'])
    # for key in dict1.keys():
    #     if key == 'Всего':
    #         # print(key, '|', dict1.get(key), dick.get(key), '|', dict1.get(key) - dick.get(key))
    #         dict1.update({key: dict1.get(key) - (dict1.get(key) - dick.get(key))})
    #         # print(key, '|', dict1.get(key), dick.get(key), '|', dict1.get(key) - dick.get(key))
    #         # print('-----------------------------------------')

    # print("Updated dick:", dick)
    # print("Updated dict1:", dict1)
    # print("Updated dict2:", dict2)

    return dict1, dict2


if __name__ == '__main__':

    sql_create_table()

    checked = False

    for branch in os.listdir(r'\\172.16.8.87\d\.rpa\.agent\robot-1p\Output\Для стата'):
        if '~' not in branch:
            print(branch)

            # if 'АФ №14' in branch or 'АФ №22' in branch or 'АФ №1' in branch or 'АФ №10' in branch or 'АСФ №3' in branch or 'АСФ №1' in branch or 'АСФ №2' in branch or 'АСФ №15' in branch or 'АФ №21' in branch:
            #     continue
            branch__ = branch.replace('Торговый зал ', '').replace('_stat.xlsx', '')
            if branch__ not in ['ШФ №32', 'АСФ №1', 'АСФ №15']:
                continue
            first = get_first_page(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-1p\Output\Для стата', branch))
            second = get_second_page(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-1p\Output\Для стата', branch))
            # print(second.keys(), second.get(list(second.keys())[0]))
            print(first)
            print(second)

            s1, s2 = 0, 0
            for key, val in first.items():
                if key != 'Всего':
                    s1 += val
            for key, val in second.items():
                if key != 'Всего':
                    s2 += val[1]
            print(s1, s2)
            first, second = get_calculated_dicts(first, second)

            print()

            print(first)
            print(second)
            s1, s2 = 0, 0
            for key, val in first.items():
                if key != 'Всего':
                    s1 += val
            for key, val in second.items():
                if key != 'Всего':
                    s2 += val[1]
            print(s1, s2)
            # sleep(1000)
            branch_ = branch.replace('_stat.xlsx', '')
            start_time = time.time()
            insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch_,
                              executor_name=ip_address, status_='processing', error_reason='', error_saved_path='', execution_time='', ecp_path_=os.path.join(ecp_paths, branch_))
            if True:
                status, error_saved_path, error = start_single_branch(os.path.join(ecp_paths, branch_), branch_, first, second)
                # status, error_saved_path, error = 'success', '', ''
                insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch_,
                                  executor_name=ip_address, status_=status, error_reason=error, error_saved_path=error_saved_path, execution_time=round(time.time() - start_time), ecp_path_=os.path.join(ecp_paths, branch_))

            # except Exception as error:
            #
            #     insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch_,
            #                       executor_name=ip_address, status_='failed with error', error_reason=str(error), error_saved_path='', execution_time=round(time.time() - start_time), ecp_path_=os.path.join(ecp_paths, branch_))


